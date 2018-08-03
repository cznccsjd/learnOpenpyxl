#coding:utf-8

"""
openpyxl文档，Simple Usage章节
url：https://openpyxl.readthedocs.io/en/stable/usage.html
"""

import time

'''
工作簿-写操作
'''
'''
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = './empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):    #  插入39行，1-39
    ws1.append(range(600))      #每行插入600列，0-599

ws2 = wb.create_sheet(title = "Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title = "Data")
for row in range(10, 20):   #插入行，10-19行
    for col in range(27, 54):   #插入列，27-53列
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))   #啥意思？
print(ws3["AA10"].value)

wb.save(filename = dest_filename)
'''

'''
工作簿-读操作
'''
'''
from openpyxl import load_workbook

wb = load_workbook(filename = './empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)
'''

'''
使用单元格格式
'''
'''
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(filename = './empty_book.xlsx')
ws = wb['Data']
# set date using a Python datetime
ws['A1'] = datetime.datetime(2018, 7, 31)
print(ws['A1'].value)
print(ws['A1'].number_format)   #打印格式
# You can enable type inference on a case-by-case basis
wb.guess_types = True
# set percentage using a string followed by the percent sign
ws['B1'] = '3.14%'
wb.guess_types = False
print(ws['B1'].value)
print(ws['B1'].number_format)
'''

'''
使用公式
'''
'''
from openpyxl import load_workbook
from openpyxl.utils import FORMULAE

wb = load_workbook(filename = './empty_book.xlsx')
ws = wb['Data']
#add a simple formula
ws['A1'] = "=SUM(1,1)"
print(ws['A1'].value)
#判断输入的内容是不是公式
print("HEX2DEC" in FORMULAE)
print("ERROR" in FORMULAE)  #随便写段字符串，肯定不在公式中

#如果你想用一个不知道的公式可能是因为你使用的公式没有包含在初始规范中。这样的公式必须以_xlfn为前缀。去工作。
'''

'''
合并/取消合并单元格
'''
'''
from openpyxl import load_workbook
from openpyxl import Workbook
# wb = load_workbook(filename='./empty_book.xlsx')
# ws = wb['Data']
wb = Workbook()
ws = wb.active

ws.merge_cells('A2:D2')
# ws.unmerge_cells('A2:D2')


#or equivalently
ws.merge_cells(start_row=2, start_column=1,end_row=2, end_column=4)
# ws.unmerge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
'''


'''
插入图片
'''
'''
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

wb = load_workbook(filename='./empty_book.xlsx')
ws = wb['Data']
ws['A1'] = 'You should see three logos below'

#create an image
img = Image('./logo.png')

#add to worksheet and anchor next to cells
ws.add_image(img, 'A1')
'''


'''
Fold columns(outline)
'''
import openpyxl
wb = openpyxl.Workbook()
ws = wb.create_sheet(title='分组')
ws.column_dimensions.group('A', 'D', hidden=True)



wb.save('./empty_book.xlsx')
# time.sleep(10)
# wb.close()