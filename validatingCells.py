#coding:utf-8
"""
openpyxl模块学习 验证单元格 Validating Cells
"""
# 数据验证器可以应用于单元格范围，但不强制或评估。范围不必是连续的:
# 例如。“A1 B2:B5”包含A1和细胞B2到B5，而不是A2或B2。
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Create the workbook and worksheet we'll be working with
wb = Workbook()
ws = wb.active

# Create a data-validation object with list validation
dv = DataValidation(type='list',formula1='"Dog,Cat,Bat"',allow_blank=True)

# Optionally set a custom error message
dv.error = 'Your entyr is not in the list'
dv.errorTitle = 'Invalid Entyr'

# Optionally set a custom prompt message
dv.prompt = 'please select from the list'
dv.promptTitle = 'List Selection'

# Add the data-validation object to the worksheet
ws.add_data_validation(dv)



# Create some cells, and add them to the data-validation object
c1 = ws["A1"]
c1.value = "Dog"
dv.add(c1)
c2 = ws["A2"]
c2.value = "An invalid value"
dv.add(c2)

# Or, apply the validation to a range of cells
dv.add('B1:B1048576')   #This is the same as for the whole of column B

#Check with a cell is in the validator
print("B4" in dv)
# 在保存工作簿时，不包含任何单元格范围的验证将被忽略。