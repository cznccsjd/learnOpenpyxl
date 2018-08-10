#coding:utf-8
"""
openpyxl学习文档：Working with Pandas and Numpy
url:http://openpyxl.readthedocs.io/en/latest/pandas.html#numpy-support
"""
### Working with Pandas Dataframes
# The openpyxl.utils.dataframe.dataframe_to_rows() function provides a simple way to work with Pandas Dataframes:
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame
from openpyxl.workbook import Workbook

wb = Workbook()
ws = wb.active

file = "./pandas.xlsx"
data = {'name':{'one':'zhangsan','two':'lisi'},'sex':{'one':'man','two':'women'}}
df = DataFrame(data)

print('\n################################################')
# for r in dataframe_to_rows(df, index=True, header=True):
#     ws.append(r)
#
# wb.save(filename=file)
# wb.close()


# 将dataframe转换为突出显示标题和索引的工作表
# To convert a dataframe into a worksheet highlighting the header and index:
'''
print('\n################################################')
for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)
# 下面代码，是excel表格第A列和第1行，加粗、加黑效果
for cell in ws['A'] + ws['1']:
    cell.style = 'Pandas'
wb.save(file)
'''

# 如果只想转换数据，可以使用只写模式
# Alternatively, if you just want to convert the data you can use write-only mode:
# print('\n################################################')
# from openpyxl.cell.cell import WriteOnlyCell
# wb = Workbook(write_only=True)
# ws = wb.create_sheet()
#
# cell = WriteOnlyCell(ws)
# cell.style = 'Pandas'
#
# def format_first_row(row, cell):
#     for c in row:
#         cell.value = c
#         yield cell
#
# rows = dataframe_to_rows(df)
# first_row = format_first_row(next(rows), cell)
# ws.append(first_row)
#
# for row in rows:
#     row = list(row)
#     cell.value = row[0]
#     row[0] = cell
#     ws.append(row)
#
# wb.save(file)
# 上面的代码与标准工作簿效果一样


