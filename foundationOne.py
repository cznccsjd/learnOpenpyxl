#coding:utf-8
"""
从excel表格中，读取数据，筛选出重复的数据

"""
import openpyxl
from openpyxl import Workbook
from locale import *


# file = r'D:\Documents\work\工作记录\180523 缺席代课报表\teacher.xlsx'
# # 打开excel
# wb = openpyxl.load_workbook(file,guess_types=True)
# # 获取当前的表格，判断是否有sheet页：teacher_info_new，如果有，不处理文档，如果没有，处理文档
# wss = wb.worksheets
# if 'teacher_info_new' not in wss:
#     wb.create_sheet('teacher_info_new')
# else:
#     pass


# # 获取当前活跃的worksheet
# ws = wb.active
# print(ws)
# print(ws)
# ws1 = wb.get_active_sheet
# print(ws1)
# ws2 = wb.active(value = '<Worksheet "teacher_info_new">')
# print(ws2)
# 获取表格最大行
# row_max = ws.max_row
# print(type(row_max))
# print('最大行：',row_max)

# 获取单个表格的值
# 直接访问
# cell = ws['A4'].value
# print(cell)
# 通过行和列获取单元格
# d = ws.cell(row = 3, column = 4).value
# print(d)


# 获取大量单元格
# 使用切片访问指定范围的单元格
# 遍历出每个单元格，再获取每个单元格的值
# cell_range = ws['A1':'B10']
# print(cell_range)
# for cell in cell_range:
#     for i in range(len(cell)):
#         print(cell[i].value)

# 访问指定行或列的单元格的值
# 获取指定列的值
# colC = ws['C']
# print(colC)
# for c in colC:
#     print(c.value)
# # 获取指定行的值
# row10 = ws[10]
# print(row10)
# for row in row10:
#     print(row.value)
# # 获取多列的值
# col_range = ws['C:D']
# print(col_range)
# for col in col_range:
#     for i in range(len(col)):
#         print(col[i].value)
# # 获取多行的值
# row_range = ws[5:10]
# print(row_range)
# for row in row_range:
#     for i in range(len(row)):
#         print(row[i].value)
#     print('\n')

# #openpyxl.worksheet.Worksheet.iter_rows() method:
# for cells in ws.iter_rows(min_col=1, max_col=3, min_row=1, max_row=3):
#     # print(cells)
#     for i in range(len(cells)):
#         print(cells[i].value)
#     print('\n')

# # 获取sheet所有的单元格
# cells = ws.rows
# for cell in cells:
#     # print(cell)
#     for i in range(len(cell)):
#         print(cell[i].value)
#     print('\n')

#### 数据存储
# print(ws['A1210'].value)
# # ws['A1210'].value = 'hello,world'
# # print(ws['A1210'].value)
#
# ws['A1210'].value = 3.14
# print(ws['A1210'].value)
# #设置单元格的格式类型
# ws['A1210'].value = '12%'
# print(ws['A1210'].value)
#
# import datetime
# ws['A1211'].value = datetime.datetime.now()
# print(ws['A1211'].value)


### 保存文件    ###
# wb = openpyxl.Workbook()
# wb.save('D:\Desktop\helo.xlsx')


# 关闭workbook对象
wb.close()
