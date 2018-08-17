#coding:utf-8
"""
openpyxl模块暂时没有看到对应的拷贝模块介绍，自己从网上找的相关信息，拷贝大体逻辑是，获取要拷贝的单元格的值，然后再保存文件（新的worksheet或者新的workbook）；
"""
from openpyxl import Workbook

dest_filename = './empty_book.xlsx'

wb = Workbook()
ws = wb.active
# 造数据
ws['A1'] = 'a1'
ws['A2'] = 'a2'

# 将数据拷贝到新的sheet页
newSheet = 'copy'
wb.create_sheet(title=newSheet)
ws1 = wb[newSheet]
cell1 = ws['A1'].value
ws1['C1'] = cell1


# 将数据拷贝到新的workbook
wb1 = Workbook()
ws11 = wb1.active
ws11['A1'] = cell1
wb1.save(filename='copy.xlsx')






#保存数据
wb.save(dest_filename)
wb.close()